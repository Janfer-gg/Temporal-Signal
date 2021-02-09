import openpyxl
from openpyxl.styles import Font, colors, Alignment,Border,Side
from Bio.Seq import MutableSeq
from math import log10

def revcomplement(seq):
    seq = seq[::-1]
    seq_dict={'A':'T','T':'A','G':'C','C':'G'}
    complement_seq=''
    for everychar in seq:
        char = seq_dict[everychar]
        complement_seq += char
    return complement_seq

def TM(seq):
    seq = MutableSeq(seq)
    GC = seq.count('G')+seq.count('C')+seq.count('g')+seq.count('c')
    tm = 100.5 + (41 * GC / len(seq)) - (820/len(seq)) + 16.6 * log10(0.05)
    return tm

def carrier_build(type,access,carname,number,gRNA1,gRNA2,scaffold,account,filepath1,filepath2):

    wb = openpyxl.load_workbook(filepath2)

    #single gRNA
    if type == "single gRNA":
        sheet1=wb.worksheets[0]
        title=["协议号","载体名称","内部编号","gRNA序列","骨架信息","Oligo-F","Oligo-F","Oligo-R","Oligo-R","检菌引物","测序引物","酶切方案"]
        result=[title]
        scaffold_name = []
        for name in sheet1['B']:
            scaffold_name.append(name.value)
        for i in range(0,account):
            Oligo_F = 'CACCg'+gRNA1[i]
            Oligo_R = 'AAAC' + revcomplement(gRNA1[i])+'c'
            content=[access[i],carname[i],number[i],gRNA1[i],scaffold[i],number[i]+'-F',Oligo_F,number[i]+'-R',Oligo_R]

            if not scaffold[i] in scaffold_name:
                return False

            for j in sheet1.rows:
                if j[1].value == scaffold[i]:
                    content.append(j[6].value)
                    content.append(j[7].value)
                    content.append(j[8].value)
            result.append(content)

        ws = openpyxl.Workbook()
        wt = ws.active

        for item in result:
            wt.append(item)

        wt.merge_cells('F1:G1')
        wt.merge_cells('H1:I1')
        wt.column_dimensions['A'].width = 20
        wt.column_dimensions['B'].width = 32
        wt.column_dimensions['C'].width = 12
        wt.column_dimensions['D'].width = 28
        wt.column_dimensions['E'].width = 12
        wt.column_dimensions['F'].width = 16
        wt.column_dimensions['G'].width = 32
        wt.column_dimensions['H'].width = 16
        wt.column_dimensions['I'].width = 32
        wt.column_dimensions['J'].width = 30
        wt.column_dimensions['K'].width = 14
        wt.column_dimensions['L'].width = 38
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        for l in range(1,wt.max_row+1):
            wt.row_dimensions[l].height = 24
            for cell in wt[l]:
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.font = Font(name='Calibri', size=11)
        for y in range(2, wt.max_row+1):
            wt[('G{}').format(y)].alignment = Alignment(vertical='center',horizontal='left')
            wt[('I{}').format(y)].alignment = Alignment(vertical='center',horizontal='left')
        ws.save('{}/carrier.xlsx'.format(filepath1))


    #dual gRNA
    if type == "dual gRNA":
        sheet1=wb.worksheets[1]
        title=["协议号", "载体名称", "内部编号", "gRNA1序列", "gRNA2序列", "骨架信息", "扩增轮次", "Oligo-F", "Oligo-F", "Oligo-R", "Oligo-R", "PCR模板", "片段名称(大小)", "检菌引物", "测序引物", "酶切方案"]
        result = [title]
        scaffold_name = []
        for name in sheet1['B']:
            scaffold_name.append(name.value)

        for i in range(0, account):

            Oligo_F1 = 'g' + gRNA1[i] + 'GTTTTAGAGCTAGAAATAGCAAGTT'
            Oligo_R1 = revcomplement(gRNA2[i]) + 'cGGTGTTTCGTCCTTTCCACAAG'

            seq1 = 'g' + gRNA1[i] + 'gttttaga'
            seq2 = revcomplement(gRNA2[i]) + 'cggtgtttc'
            m=n=21

            while TM(seq1[0:m])>70:
                m = m - 1
                if TM(seq1[0:m])<=70 or m == 18:
                    break
            while TM(seq1[0:m])<60:
                m =m + 1
                if TM(seq1[0:m])>=60 or m == 29:
                    break

            while TM(seq2[0:n])>70:
                n = n - 1
                if TM(seq2[0:n])<=70 or n == 18:
                    break
            while TM(seq2[0:n])<60:
                n =n + 1
                if TM(seq2[0:n])>=60 or n == 29:
                    break

            while TM(seq1[0:m]) - TM(seq2[0:n]) >6:
                if n<29:
                    n= n + 1
                    if TM(seq1[0:m]) - TM(seq2[0:n]) <= 6:
                        break
                if m>18:
                    m= m - 1
                    if TM(seq1[0:m]) - TM(seq2[0:n]) <= 6:
                        break
                if n==29 and m==18:
                    break

            while TM(seq2[0:n]) - TM(seq1[0:m]) >6:
                if m<29:
                    m = m + 1
                    if TM(seq2[0:n]) - TM(seq1[0:m]) <=6:
                        break
                if n>18:
                    n = n - 1
                    if TM(seq2[0:n]) - TM(seq1[0:m]) <=6:
                        break
                if m==29 and n==18:
                    break

            Oligo_F2 = 'TGTGGAAAGGACGAAACACC' + seq1[0:m]
            Oligo_R2 = 'GCTATTTCTAGCTCTAAAAC' + seq2[0:n]
            content1 = [access[i], carname[i], number[i], gRNA1[i], gRNA2[i], scaffold[i], '一轮PCR', number[i] + '-F1', Oligo_F1, number[i] + '-R1', Oligo_R1,"合成基因scaffold-U6",number[i]+"-1 (381bp)"]
            content2 = [access[i], carname[i], number[i], gRNA1[i], gRNA2[i], scaffold[i], '二轮PCR', number[i] + '-F2', Oligo_F2, number[i] + '-R2', Oligo_R2,number[i]+"-1",number[i]+"-2 (420bp)"]

            if not scaffold[i] in scaffold_name:
                return False

            for j in sheet1.rows:
                if j[1].value == scaffold[i]:
                    content1.append(j[9].value)
                    content1.append(j[10].value)
                    content1.append(j[11].value)

            result.append(content1)
            result.append(content2)

        ws = openpyxl.Workbook()
        wt = ws.active

        for item in result:
            wt.append(item)


        wt.merge_cells('H1:I1')
        wt.merge_cells('J1:K1')
        for i in range(1,account+1):
            wt.merge_cells(('A{}:A{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('B{}:B{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('C{}:C{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('D{}:D{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('E{}:E{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('F{}:F{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('N{}:N{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('O{}:O{}').format(i * 2, i * 2 + 1))
            wt.merge_cells(('P{}:P{}').format(i * 2, i * 2 + 1))

        wt.column_dimensions['A'].width = 20
        wt.column_dimensions['B'].width = 36
        wt.column_dimensions['C'].width = 12
        wt.column_dimensions['D'].width = 28
        wt.column_dimensions['E'].width = 28
        wt.column_dimensions['F'].width = 12
        wt.column_dimensions['G'].width = 10
        wt.column_dimensions['H'].width = 16
        wt.column_dimensions['I'].width = 32
        wt.column_dimensions['J'].width = 16
        wt.column_dimensions['K'].width = 32
        wt.column_dimensions['L'].width = 20
        wt.column_dimensions['M'].width = 20
        wt.column_dimensions['N'].width = 30
        wt.column_dimensions['O'].width = 24
        wt.column_dimensions['P'].width = 38

        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        for l in range(1, wt.max_row + 1):
            wt.row_dimensions[l].height = 24
            for cell in wt[l]:
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.font = Font(name='Calibri', size=11)

        for y in range(2, wt.max_row+1):
            wt[('I{}').format(y)].alignment = Alignment(vertical='center',horizontal='left')
            wt[('k{}').format(y)].alignment = Alignment(vertical='center',horizontal='left')

        ws.save('{}/carrier.xlsx'.format(filepath1))

    #shRNA
    if type == "shRNA":
        sheet1 = wb.worksheets[2]
        title = ["协议号", "载体名称", "内部编号", "shRNA序列", "骨架信息", "Oligo-F", "Oligo-F", "Oligo-R", "Oligo-R", "检菌引物", "测序引物","酶切方案"]
        result = [title]
        scaffold_name = []
        for name in sheet1['B']:
            scaffold_name.append(name.value)

        for i in range(0, account):
            Oligo_F = 'CCGG' + gRNA1[i] + 'CTCGAG' + revcomplement(gRNA1[i]) +'TTTTTG'
            Oligo_R = 'AATTCAAAAA' + gRNA1[i] + 'CTCGAG' + revcomplement(gRNA1[i])
            content = [access[i], carname[i], number[i], gRNA1[i], scaffold[i], number[i] + '-F', Oligo_F, number[i] + '-R', Oligo_R]

            if not scaffold[i] in scaffold_name:
                return False

            for j in sheet1.rows:
                if j[1].value == scaffold[i]:
                    content.append(j[6].value)
                    content.append(j[7].value)
                    content.append(j[8].value)
            result.append(content)


        ws = openpyxl.Workbook()
        wt = ws.active

        for item in result:
            wt.append(item)

        wt.merge_cells('F1:G1')
        wt.merge_cells('H1:I1')
        wt.column_dimensions['A'].width = 20
        wt.column_dimensions['B'].width = 32
        wt.column_dimensions['C'].width = 12
        wt.column_dimensions['D'].width = 28
        wt.column_dimensions['E'].width = 12
        wt.column_dimensions['F'].width = 16
        wt.column_dimensions['G'].width = 32
        wt.column_dimensions['H'].width = 16
        wt.column_dimensions['I'].width = 32
        wt.column_dimensions['J'].width = 30
        wt.column_dimensions['K'].width = 20
        wt.column_dimensions['L'].width = 38
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        for l in range(1, wt.max_row + 1):
            wt.row_dimensions[l].height = 24
            for cell in wt[l]:
                cell.alignment = Alignment(vertical='center', horizontal='center')
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.font = Font(name='Calibri', size=11)
        for y in range(2, wt.max_row + 1):
            wt[('G{}').format(y)].alignment = Alignment(vertical='center', horizontal='left')
            wt[('I{}').format(y)].alignment = Alignment(vertical='center', horizontal='left')
        ws.save('{}/carrier.xlsx'.format(filepath1))


# if __name__ == '__main__':
#     type ="dual gRNA"
#     access =['UBISCM201020WL1','UBISCM201020WL1']
#     carname = ['YKO-RP003-BRAF(c.T1799A)[gRNA1]','YKO-RP003-BRAF(c.T1799A)[gRNA2]']
#     number = ['CM20-015a','CM20-015b']
#     gRNA1=['TAGCTACAGTGAAATCTCGA','ACAGTGAAATCTCGATGGAG']
#     gRNA2 = ['ACAGTGAAATCTCGATGGAG','TAGCTACAGTGAAATCTCGA']
#     scaffold = ['YKO-RP003','YKO-RP003']
#     account=2
#     carrier_build(type,access,carname,number,gRNA1,gRNA2,scaffold,account,'','')
